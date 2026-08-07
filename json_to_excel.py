#!/usr/bin/env python3
"""Convert JSON files into a formatted Excel workbook, one tab per file.

This is the RAW converter: no filtering, no scoring, no history. Use it to
eyeball a JSON file, hand someone unabridged data, or work out why a posting
never reached the Shortlist tab.

    python json_to_excel.py                       # data/postings.json
    python json_to_excel.py --all                 # every .json in data/, one tab each
    python json_to_excel.py path/to/file.json     # a specific file
    python json_to_excel.py -o report.xlsx        # choose the output path
    python json_to_excel.py --sort score --desc   # sort by a column
    python json_to_excel.py --no-summary          # skip the Summary tab

`run.py` is the thing you normally want -- it collects, filters, scores, and
writes the six-tab workbook. This is the escape hatch. It also handles arbitrary
JSON, so it works on data that never came from this project.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from src.export import style

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
OUT_DIR = ROOT / "output"

# Re-exported so tests and callers have one obvious import site.
clean = style.clean
order_columns = style.order_columns
sort_rows = style.sort_rows


def load_rows(path: Path) -> list[dict]:
    """Accept a bare list, a dict wrapping one, or a single object."""
    doc = json.loads(path.read_text(encoding="utf-8"))

    if isinstance(doc, dict):
        for key in ("postings", "jobs", "results", "items", "data"):
            if isinstance(doc.get(key), list):
                doc = doc[key]
                break
        else:
            # history.json is keyed by posting id -> record
            values = list(doc.values())
            doc = values if values and all(isinstance(v, dict) for v in values) else [doc]

    if not isinstance(doc, list):
        raise ValueError(
            f"{path.name}: expected a list of objects, got {type(doc).__name__}"
        )

    rows = [r for r in doc if isinstance(r, dict)]
    if len(rows) != len(doc):
        print(f"  ! skipped {len(doc) - len(rows)} non-object entries", file=sys.stderr)
    return rows


def counts(rows: list[dict], field: str) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in rows:
        if field in r:
            k = str(r.get(field) or "(blank)")
            out[k] = out.get(k, 0) + 1
    return out


def convert(paths: list[Path], out_path: Path, sort: str | None = None,
            desc: bool = False, summary: bool = True) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    all_rows: list[dict] = []
    for p in paths:
        rows = load_rows(p)
        if not rows:
            print(f"  ! {p.name} has no rows, skipping", file=sys.stderr)
            continue
        if sort:
            if not any(sort in r for r in rows):
                print(f"  ! no column {sort!r} in {p.name}, leaving order as-is",
                      file=sys.stderr)
            else:
                rows = sort_rows(rows, sort, desc)

        style.write_table(
            wb.create_sheet(p.stem[:31]),
            rows,
            title=f"{p.name} — {len(rows)} rows — exported {date.today().isoformat()}",
            subtitle="Raw export: no filtering, no scoring. Every row in the source file.",
        )
        all_rows.extend(rows)
        print(f"  {p.name}: {len(rows)} rows -> tab {p.stem[:31]!r}")

    if not wb.sheetnames:
        raise SystemExit("nothing to write — no usable rows found")

    if summary and all_rows:
        style.write_counts(
            wb.create_sheet("Summary"), "Summary",
            [(f"By {f}", counts(all_rows, f))
             for f in ("institution", "platform", "state", "sponsorship_flag", "status")],
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("source", nargs="?", default=None,
                    help="JSON file (default: data/postings.json)")
    ap.add_argument("--all", action="store_true",
                    help="convert every .json in data/, one tab each")
    ap.add_argument("-o", "--out", default=None, help="output .xlsx path")
    ap.add_argument("--sort", default=None, help="column to sort by, e.g. score")
    ap.add_argument("--desc", action="store_true", help="sort descending")
    ap.add_argument("--no-summary", action="store_true", help="skip the Summary tab")
    args = ap.parse_args()

    if args.all:
        paths = sorted(DATA_DIR.glob("*.json"))
        if not paths:
            raise SystemExit(f"no .json files in {DATA_DIR}")
    else:
        p = Path(args.source) if args.source else DATA_DIR / "postings.json"
        if not p.exists():
            raise SystemExit(f"not found: {p}")
        paths = [p]

    stem = "data_export" if args.all else paths[0].stem
    out = Path(args.out) if args.out else OUT_DIR / f"{stem}_{date.today().isoformat()}.xlsx"

    print(f"converting {len(paths)} file(s)")
    written = convert(paths, out, args.sort, args.desc, not args.no_summary)
    print(f"wrote {written}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
