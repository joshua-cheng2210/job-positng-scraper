"""The one workbook a run produces.

Tabs
    Shortlist            current run, filtered + scored, full-time only, best first.
                         Hard-blockered postings (citizenship, clearance, export
                         control, student-status -- see filters.py's HARD_BLOCKERS)
                         are excluded entirely, not just flagged.
    Part-Time & Temporary  same filtering as Shortlist, but for postings tagged
                         part-time or temporary -- kept off the main Shortlist so
                         they don't compete with full-time roles, not because
                         they're worse.
    Institutions         Shortlist (full-time only) grouped by institution,
                         ranked by a composite score
    All Postings         current run, unfiltered
    History               every posting ever collected, deduped, open + closed
    Changes               new and closed since the previous run
    Summary               counts by institution / portal / state / sponsorship / AI Score
    Run Stats             what the filters dropped and why

Filename is workbook_<date>_<time>.xlsx, so runs never overwrite each other.
`prune()` caps how many of these pile up in output/ -- run.py calls it after
every write to keep only the most recent 5.

AI Score / AI Score Reason columns are merged in from data/ai_scores.json
(see ai_score_key() below) if that file exists -- a durable side-car the
/ai-score-shortlist skill writes to, since the LLM judgment behind them is
too expensive to redo every run and a fresh workbook can't be the source of
truth for it. Blank AI Score means not yet scored, not "scored zero."
"""
from __future__ import annotations

import logging
import math
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from ..models import Posting
from . import style

log = logging.getLogger(__name__)

BULKY = {"description", "raw"}          # too wide for the overview tabs -- dropped entirely

# `url` is dropped too, but for a different reason: style.write_table() now
# hyperlinks the Title cell straight to the posting instead, so a separate
# URL column is redundant. `url` stays in the row dict (read directly by
# write_table to build that hyperlink) -- only the rendered column goes away.
DROP_COLUMNS = BULKY | {"url"}

# Kept in the data (Summary counts, etc. can still read them) but collapsed
# by default so the table isn't cluttered -- rarely-populated detail fields
# most useful when hunting down a specific row. State used to be folded into
# a combined Location column and hidden too; reverted back to its own
# visible column per Josh's call.
HIDDEN_COLUMNS = {"department", "sponsorship_evidence", "hard_blockers"}


def ai_score_key(institution: str | None, job_id: str | None) -> str:
    """Same join key data/history.json already uses (institution|job_id,
    lowercased) -- so data/ai_scores.json lines up with both history.json
    and postings.json without a third key scheme to keep in sync. Not
    hashed (unlike Posting.key) since this key needs to be human-writable/
    -readable in a JSON file a human (or the /ai-score-shortlist skill) is
    directly reading and editing."""
    return f"{(institution or '').lower()}|{(job_id or '').lower()}"


def _prepare_rows(rows: list[dict], ai_scores: dict[str, dict] | None = None) -> list[dict]:
    """Row-level display transforms shared by every table tab:

    - days_since_posted is computed fresh at export time from posted_date,
      since "how long ago" only makes sense relative to when the workbook
      was written, not when the posting was scraped.
    - ai_score/ai_reason are merged in from data/ai_scores.json (see
      ai_score_key above) if provided -- that file is the durable side-car
      the /ai-score-shortlist skill writes to, since a fresh workbook is
      generated every run and can't be the source of truth for something
      that's expensive (an LLM judgment) to redo.
    """
    today = date.today()
    out = []
    for r in rows:
        r = dict(r)
        pd = r.get("posted_date")
        days = None
        if pd:
            try:
                d = date.fromisoformat(pd) if isinstance(pd, str) else pd
                days = (today - d).days
            except (TypeError, ValueError):
                days = None
        r["days_since_posted"] = days

        if ai_scores:
            hit = ai_scores.get(ai_score_key(r.get("institution"), r.get("job_id")))
            if hit:
                r["ai_score"] = hit.get("ai_score")
                r["ai_reason"] = hit.get("ai_reason")

        out.append(r)
    return out


def _counts(rows: list[dict], field: str) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in rows:
        k = str(r.get(field) or "(blank)")
        out[k] = out.get(k, 0) + 1
    return out


def _ai_score_counts(rows: list[dict]) -> dict[str, int]:
    """Buckets by the exact 0-10 AI Score, plus a "(not yet AI-scored)"
    bucket -- seeing that count shrink run over run is the signal that
    /ai-score-shortlist is actually keeping up with the Shortlist."""
    out: dict[str, int] = {}
    for r in rows:
        v = r.get("ai_score")
        if v is None:
            k = "(not yet AI-scored)"
        elif int(v) == 0:
            k = "0 (not recommended)"
        else:
            k = str(int(v))
        out[k] = out.get(k, 0) + 1
    return out


def institution_rankings(rows: list[dict]) -> list[dict]:
    """Group Shortlist rows by institution and rank by a composite that
    rewards both QUALITY (top-3 average score) and QUANTITY (more postings,
    diminishing after 5 -- via sqrt so a 6th or 60th posting barely moves the
    needle once a school has already proven it has real openings). This is
    deliberate: a school with one lucky 11.5 and nothing else shouldn't
    outrank a school with five solid 7s.

    composite = top3_avg_score * sqrt(min(postings, 5))

    verified_pct matters as much as the score -- see Posting.description_scraped.
    A school at 0% verified means none of its scores are backed by a real
    fetched description yet; don't prioritize it over a lower-scoring but
    fully-verified school until enrichment actually runs for it.
    """
    by_inst: dict[str, list[dict]] = {}
    for r in rows:
        by_inst.setdefault(r.get("institution") or "(unknown)", []).append(r)

    out: list[dict] = []
    for inst, group in by_inst.items():
        scores = [float(r.get("score") or 0.0) for r in group]
        n = len(group)
        top3 = sorted(scores, reverse=True)[:3]
        top3_avg = sum(top3) / len(top3) if top3 else 0.0
        verified = sum(1 for r in group if r.get("description_scraped") == 1)

        out.append({
            "institution": inst,
            "postings": n,
            "max_score": round(max(scores), 2) if scores else 0.0,
            "avg_score": round(sum(scores) / n, 2) if n else 0.0,
            "top3_avg_score": round(top3_avg, 2),
            "verified_pct": round(verified / n * 100) if n else 0,
            "positive_sponsorship": sum(1 for r in group
                                        if r.get("sponsorship_flag") == "h1b_possible"),
            "no_sponsorship": sum(1 for r in group
                                  if r.get("sponsorship_flag") == "no_sponsorship_any"),
            "composite": round(top3_avg * math.sqrt(min(n, 5)), 2),
        })

    out.sort(key=lambda r: r["composite"], reverse=True)
    return out


def filename(out_dir: str | Path, when: datetime | None = None) -> Path:
    when = when or datetime.now()
    return Path(out_dir) / f"workbook_{when:%Y-%m-%d_%H%M%S}.xlsx"


def prune(out_dir: str | Path, keep: int = 5, dry_run: bool = False) -> list[Path]:
    """Delete all but the `keep` most recent workbook_*.xlsx files in out_dir.

    The timestamp is in the filename (workbook_<date>_<time>.xlsx), so a plain
    reverse-alphabetical sort is also a reverse-chronological sort -- no stat()
    call needed. Returns the paths that were (or, with dry_run=True, would be)
    deleted, so the caller can log them; deletion failures (e.g. a file open
    in Excel) are logged and skipped rather than raised, so pruning never
    fails a run.
    """
    out_dir = Path(out_dir)
    books = sorted(out_dir.glob("workbook_*.xlsx"), reverse=True)
    doomed = books[keep:]
    if dry_run:
        return doomed

    deleted: list[Path] = []
    for book in doomed:
        try:
            book.unlink()
            deleted.append(book)
        except OSError as exc:
            log.warning("could not delete %s: %s", book, exc)
    return deleted


def write(out_path: str | Path, *, shortlist: list[Posting], all_postings: list[Posting],
          history_rows: list[dict], changes: dict[str, list[dict]],
          part_time: list[Posting] | None = None,
          stats: dict[str, int] | None = None,
          ai_scores: dict[str, dict] | None = None) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    short_rows = _prepare_rows([p.to_row() for p in shortlist], ai_scores)
    part_time_rows = _prepare_rows([p.to_row() for p in (part_time or [])], ai_scores)
    all_rows = _prepare_rows([p.to_row() for p in all_postings], ai_scores)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()
    wb.remove(wb.active)

    style.write_table(
        wb.create_sheet("Shortlist"),
        style.sort_rows(short_rows, "ai_score", desc=True),
        title=f"Shortlist — {len(short_rows)} postings — {stamp}",
        subtitle=("Survived the filters and carries no hard blocker (citizenship, "
                  "clearance, export control, student-status), sorted by AI Score "
                  "(postings not yet AI-scored sink to the bottom -- run "
                  "/ai-score-shortlist to score them, or see the Score column for the "
                  "regex-based fallback ranking). Full-time only -- part-time and "
                  "temporary postings are on their own tab. Green = positive "
                  "sponsorship language, amber = declines STEM OPT (still applicable "
                  "via cap-exempt H-1B), red = no sponsorship."),
        drop=DROP_COLUMNS,
        hidden=HIDDEN_COLUMNS,
    )

    style.write_table(
        wb.create_sheet("Part-Time & Temporary"),
        style.sort_rows(part_time_rows, "ai_score", desc=True),
        title=f"Part-Time & Temporary — {len(part_time_rows)} postings — {stamp}",
        subtitle=("Same filtering as Shortlist (survived the filters, no hard blocker) "
                  "but tagged part-time or temporary in the posting text -- split out "
                  "so these don't compete with full-time roles for ranking, not "
                  "because they're worse. Run /ai-score-shortlist to score these too."),
        drop=DROP_COLUMNS,
        hidden=HIDDEN_COLUMNS,
    )

    # Institution rankings use score/sponsorship/verified fields only, so the
    # location/state merge above doesn't matter here -- computed from the
    # already-prepared short_rows is fine either way.
    rankings = institution_rankings(short_rows)
    style.write_table(
        wb.create_sheet("Institutions"),
        rankings,
        title=f"Institution Rankings — {len(rankings)} institutions — {stamp}",
        subtitle=("Composite = top-3 average score x sqrt(min(postings, 5)) -- rewards "
                  "a school with several solid postings over one lucky outlier. "
                  "Verified % near 0 means don't trust that school's scores yet -- "
                  "enrichment hasn't confirmed the descriptions behind them."),
    )

    style.write_table(
        wb.create_sheet("All Postings"),
        all_rows,
        title=f"All postings collected — {len(all_rows)} rows — {stamp}",
        subtitle="Unfiltered. Use this to check why something never reached the Shortlist.",
        drop=DROP_COLUMNS,
        hidden=HIDDEN_COLUMNS,
    )

    style.write_table(
        wb.create_sheet("History"),
        _prepare_rows(history_rows, ai_scores),
        title=f"History — {len(history_rows)} unique postings ever seen",
        subtitle=("Carried forward across every run and deduplicated by institution + job ID. "
                  "status=closed means it was not in the latest run."),
        drop=DROP_COLUMNS,
        hidden=HIDDEN_COLUMNS,
    )

    change_rows = _prepare_rows(changes.get("new", []) + changes.get("closed", []), ai_scores)
    style.write_table(
        wb.create_sheet("Changes"),
        change_rows,
        title=f"Changes this run — {len(changes.get('new', []))} new, "
              f"{len(changes.get('closed', []))} closed",
        subtitle="Only what moved since the previous run. Empty on a first run.",
        drop=DROP_COLUMNS,
        hidden=HIDDEN_COLUMNS,
    )

    # Negative-score postings (hard blocker, no-sponsorship, 5+ years, etc.)
    # dragged the "worth applying to" breakdowns down with institutions/
    # sponsorship flags that are really just noise -- a school with three
    # postings that all net negative shouldn't show up here next to schools
    # with real prospects. Only these two Shortlist breakdowns are filtered;
    # "All postings by portal/state" and "History by status" are collection
    # health checks, not application-worthiness checks, so they stay
    # unfiltered on purpose.
    worth_applying = [r for r in short_rows if (r.get("score") or 0) >= 0]

    style.write_counts(
        wb.create_sheet("Summary"),
        f"Summary — {stamp}",
        [
            ("Shortlist by institution (score >= 0 only)",
             _counts(worth_applying, "institution")),
            ("Shortlist by sponsorship flag (score >= 0 only)",
             _counts(worth_applying, "sponsorship_flag")),
            # Each Shortlist breakdown sits directly above its All-Postings
            # counterpart so the two are easy to compare at a glance --
            # e.g. "12 of these 40 MN postings actually made the Shortlist."
            ("Shortlist by portal (score >= 0 only)",
             _counts(worth_applying, "platform")),
            ("All postings by portal", _counts(all_rows, "platform")),
            ("Shortlist by state (score >= 0 only)",
             _counts(worth_applying, "state")),
            ("All postings by state", _counts(all_rows, "state")),
            ("Shortlist by AI Score (score >= 0 only)",
             _ai_score_counts(worth_applying)),
            ("Part-Time & Temporary by AI Score", _ai_score_counts(part_time_rows)),
            ("History by status", _counts(history_rows, "status")),
        ],
    )

    if stats:
        style.write_counts(wb.create_sheet("Run Stats"), f"Run stats — {stamp}",
                           [("Filter outcome", dict(stats))])

    wb.save(out_path)
    return out_path
