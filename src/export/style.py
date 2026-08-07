"""Shared Excel styling. One place, so the tabs can't drift apart."""
from __future__ import annotations

from datetime import date, datetime
import json

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FONT = "Arial"
NAVY = "1F3864"
GREY = "F2F2F2"
GREEN = "E2EFDA"
AMBER = "FFF2CC"
RED = "FCE4E4"
BLUE = "DDEBF7"

MAX_CELL = 32_000                      # Excel's hard limit is 32,767

FLAG_FILL = {
    "h1b_possible": GREEN,
    "no_stem_opt": AMBER,
    "no_sponsorship_any": RED,
}
STATUS_FILL = {
    "new": GREEN,
    "open": None,
    "closed": RED,
}

HEADERS = {
    "score": "Score", "institution": "Institution", "title": "Title",
    "department": "Department", "location": "Location", "state": "State",
    "job_id": "Job ID", "posted_date": "Posted", "close_date": "Closes",
    "sponsorship_flag": "Sponsorship", "sponsorship_evidence": "Sponsorship evidence",
    "hard_blockers": "Blockers", "platform": "Portal", "system": "System",
    "url": "URL", "score_reasons": "Why this score", "description": "Description",
    "description_scraped": "Description Verified?",
    "first_seen": "First seen", "last_seen": "Last seen", "status": "Status",
    "runs_seen": "Runs seen", "change": "Change",
    "postings": "Postings", "max_score": "Max Score", "avg_score": "Avg Score",
    "top3_avg_score": "Top-3 Avg", "verified_pct": "Verified %",
    "positive_sponsorship": "Positive Sponsorship Count",
    "no_sponsorship": "No Sponsorship Count", "composite": "Composite Rank",
}

WIDTHS = {
    "score": 8, "institution": 34, "title": 46, "department": 28, "location": 22,
    "state": 7, "job_id": 16, "posted_date": 12, "close_date": 12,
    "sponsorship_flag": 18, "sponsorship_evidence": 34, "hard_blockers": 26,
    "platform": 13, "system": 22, "url": 58, "score_reasons": 52,
    "description": 80, "description_scraped": 12,
    "first_seen": 12, "last_seen": 12, "status": 11,
    "runs_seen": 10, "change": 12,
    "postings": 10, "max_score": 10, "avg_score": 10, "top3_avg_score": 10,
    "verified_pct": 11, "positive_sponsorship": 16, "no_sponsorship": 15,
    "composite": 13,
}

PREFERRED = [
    "change", "status", "score", "institution",
    "composite", "postings", "top3_avg_score", "avg_score", "max_score",
    "verified_pct", "positive_sponsorship", "no_sponsorship",
    "title", "department", "location",
    "state", "job_id", "posted_date", "close_date", "first_seen", "last_seen",
    "runs_seen", "sponsorship_flag", "sponsorship_evidence", "hard_blockers",
    "platform", "system", "url", "score_reasons", "description_scraped", "description",
]

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def order_columns(rows: list[dict], drop: set[str] | None = None) -> list[str]:
    seen: set[str] = set()
    for r in rows:
        seen.update(r.keys())
    seen -= (drop or set())
    ordered = [c for c in PREFERRED if c in seen]
    return ordered + sorted(seen - set(ordered))


def clean(value):
    """Coerce anything JSON can hold into something Excel accepts."""
    if value is None or isinstance(value, (int, float, bool, datetime, date)):
        return value
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(v) for v in value)
    elif isinstance(value, dict):
        value = json.dumps(value, default=str)
    else:
        value = str(value)
    if len(value) == 10 and value[4:5] == "-" and value[7:8] == "-":
        try:
            return date.fromisoformat(value)
        except ValueError:
            pass
    return value[:MAX_CELL] if len(value) > MAX_CELL else value


def sort_key(row: dict, field: str):
    """Total ordering over mixed types -- fixed-shape tuple, so int never meets str."""
    v = row.get(field)
    if isinstance(v, bool):
        return (0, float(v), "")
    if isinstance(v, (int, float)):
        return (0, float(v), "")
    if isinstance(v, (date, datetime)):
        return (0, 0.0, v.isoformat())
    return (1, 0.0, str(v).lower())


def sort_rows(rows: list[dict], field: str, desc: bool = True) -> list[dict]:
    """Sort by field; rows with no value stay at the bottom in both directions."""
    have = [r for r in rows if r.get(field) is not None]
    missing = [r for r in rows if r.get(field) is None]
    have.sort(key=lambda r: sort_key(r, field), reverse=desc)
    return have + missing


def write_table(ws: Worksheet, rows: list[dict], *, title: str, subtitle: str = "",
                drop: set[str] | None = None) -> None:
    """Render rows as a formatted, filterable table starting at row 4."""
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.sheet_view.showGridLines = False

    if not rows:
        ws["A4"] = "(nothing to show)"
        ws["A4"].font = Font(name=FONT, italic=True, size=10, color="808080")
        return

    cols = order_columns(rows, drop)
    hdr = 4
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=hdr, column=i, value=HEADERS.get(col, col))
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 26

    idx = {c: i + 1 for i, c in enumerate(cols)}

    for r, row in enumerate(rows, start=hdr + 1):
        for col in cols:
            cell = ws.cell(row=r, column=idx[col], value=clean(row.get(col)))
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(col != "url"))
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"

        if "sponsorship_flag" in idx:
            fill = FLAG_FILL.get(row.get("sponsorship_flag"))
            if fill:
                ws.cell(row=r, column=idx["sponsorship_flag"]).fill = \
                    PatternFill("solid", fgColor=fill)

        if "hard_blockers" in idx and row.get("hard_blockers"):
            ws.cell(row=r, column=idx["hard_blockers"]).fill = \
                PatternFill("solid", fgColor=RED)

        if "description_scraped" in idx and row.get("description_scraped") == 1:
            ws.cell(row=r, column=idx["description_scraped"]).fill = \
                PatternFill("solid", fgColor=GREEN)

        for field in ("status", "change"):
            if field in idx:
                fill = STATUS_FILL.get(str(row.get(field, "")).lower())
                if fill:
                    ws.cell(row=r, column=idx[field]).fill = \
                        PatternFill("solid", fgColor=fill)

        if "url" in idx:
            uc = ws.cell(row=r, column=idx["url"])
            if isinstance(uc.value, str) and uc.value.startswith("http"):
                uc.hyperlink = uc.value
                uc.font = Font(name=FONT, size=10, color="0563C1", underline="single")

    for col in cols:
        ws.column_dimensions[get_column_letter(idx[col])].width = WIDTHS.get(col, 20)
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{len(rows) + hdr}"


def write_counts(ws: Worksheet, title: str, blocks: list[tuple[str, dict[str, int]]]) -> None:
    """Stacked label/count tables on one sheet."""
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    ws.sheet_view.showGridLines = False

    r = 3
    for heading, counts in blocks:
        if not counts:
            continue
        ws.cell(row=r, column=1, value=heading).font = \
            Font(name=FONT, bold=True, size=11, color=NAVY)
        r += 1
        for col, label in ((1, "Value"), (2, "Count")):
            c = ws.cell(row=r, column=col, value=label)
            c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
        r += 1
        for k, v in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10)
            ws.cell(row=r, column=2, value=v).font = Font(name=FONT, size=10)
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 10
