"""Converter tests. Every case here came from an actual failure during build."""
import json
from datetime import date

import pytest
from openpyxl import load_workbook

from json_to_excel import clean, convert, load_rows, order_columns, sort_rows


def write(tmp_path, name, payload):
    p = tmp_path / name
    p.write_text(json.dumps(payload), encoding="utf-8")
    return p


def test_load_bare_list(tmp_path):
    p = write(tmp_path, "a.json", [{"x": 1}, {"x": 2}])
    assert len(load_rows(p)) == 2


def test_load_dict_wrapped_list(tmp_path):
    p = write(tmp_path, "b.json", {"postings": [{"x": 1}]})
    assert load_rows(p) == [{"x": 1}]


def test_load_single_object_becomes_one_row(tmp_path):
    p = write(tmp_path, "c.json", {"institution": "Solo U"})
    assert load_rows(p) == [{"institution": "Solo U"}]


def test_load_rejects_scalar(tmp_path):
    p = write(tmp_path, "d.json", "just a string")
    with pytest.raises(ValueError):
        load_rows(p)


def test_known_columns_come_first_unknown_appended_sorted():
    cols = order_columns([{"zebra": 1, "title": "t", "alpha": 2, "score": 3}])
    assert cols[:2] == ["score", "title"]
    assert cols[2:] == ["alpha", "zebra"]


def test_clean_flattens_lists_and_dicts():
    assert clean(["a", "b"]) == "a; b"
    assert json.loads(clean({"k": "v"})) == {"k": "v"}


def test_clean_converts_iso_dates():
    assert clean("2026-08-11") == date(2026, 8, 11)
    assert clean("not-a-date") == "not-a-date"


def test_clean_truncates_beyond_excel_limit():
    assert len(clean("x" * 50_000)) == 32_000


def test_sort_survives_mixed_types():
    """The bug that crashed the first version: int vs str in one column."""
    rows = [{"s": 5}, {"s": None}, {"s": "text"}, {"s": True}]
    out = sort_rows(rows, "s", desc=True)
    assert len(out) == 4
    assert out[-1]["s"] is None          # None stays last when descending


def test_sort_keeps_none_last_ascending_too():
    rows = [{"s": 5}, {"s": None}, {"s": 1}]
    out = sort_rows(rows, "s", desc=False)
    assert [r["s"] for r in out] == [1, 5, None]


def test_end_to_end_workbook(tmp_path):
    src = write(tmp_path, "postings.json", [
        {"institution": "A U", "title": "Research Software Engineer", "score": 15.5,
         "url": "https://example.edu/1", "sponsorship_flag": "h1b_possible",
         "hard_blockers": "", "posted_date": "2026-07-28", "platform": "workday"},
        {"institution": "B U", "title": "Systems Analyst", "score": -12.0,
         "url": "https://example.edu/2", "sponsorship_flag": "unknown",
         "hard_blockers": "US citizenship required", "platform": "peopleadmin"},
    ])
    out = tmp_path / "out.xlsx"
    convert([src], out, sort="score", desc=True, summary=True)

    wb = load_workbook(out)
    assert "postings" in wb.sheetnames
    assert "Summary" in wb.sheetnames

    ws = wb["postings"]
    assert ws["A4"].value == "Score"                 # header row
    assert ws["A5"].value == 15.5                    # sorted desc
    assert ws["A6"].value == -12.0
    assert ws.freeze_panes == "A5"
    assert ws.auto_filter.ref is not None

    url_col = order_columns(load_rows(src)).index("url") + 1
    assert ws.cell(row=5, column=url_col).hyperlink is not None


def test_empty_file_raises_cleanly(tmp_path):
    src = write(tmp_path, "empty.json", [])
    with pytest.raises(SystemExit):
        convert([src], tmp_path / "o.xlsx", sort=None, desc=False, summary=False)
