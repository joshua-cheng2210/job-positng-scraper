"""The seven-tab workbook."""
from datetime import date, datetime

from openpyxl import load_workbook

from src.export import workbook
from src.models import Posting


def mk(inst, jid, title, score=0.0, flag="unknown", **kw):
    p = Posting(institution=inst, job_id=jid, title=title,
                url=f"https://x/{jid}", platform="workday", state="MN", **kw)
    p.score = score
    p.sponsorship_flag = flag
    return p


def test_filename_is_timestamped(tmp_path):
    f = workbook.filename(tmp_path, datetime(2026, 8, 6, 18, 38, 3))
    assert f.name == "workbook_2026-08-06_183803.xlsx"


def test_two_runs_never_collide(tmp_path):
    a = workbook.filename(tmp_path, datetime(2026, 8, 6, 18, 0, 0))
    b = workbook.filename(tmp_path, datetime(2026, 8, 6, 18, 0, 1))
    assert a != b


def test_all_eight_tabs_present(tmp_path):
    out = tmp_path / "wb.xlsx"
    short = [mk("A U", "1", "Software Engineer", 12.0, "h1b_possible")]
    every = short + [mk("A U", "2", "Groundskeeper")]
    hist = [{"institution": "A U", "job_id": "1", "title": "Software Engineer",
             "status": "open", "first_seen": "2026-08-01", "last_seen": "2026-08-06",
             "runs_seen": 3}]
    changes = {"new": [dict(hist[0], change="new")], "closed": []}

    workbook.write(out, shortlist=short, all_postings=every,
                   history_rows=hist, changes=changes, stats={"kept": 1})

    wb = load_workbook(out)
    assert wb.sheetnames == ["Shortlist", "Part-Time & Temporary", "Institutions",
                             "All Postings", "History", "Changes", "Summary", "Run Stats"]
    assert wb["Shortlist"].freeze_panes == "A5"
    assert wb["Shortlist"].auto_filter.ref is not None


def test_part_time_tab_receives_its_own_rows(tmp_path):
    out = tmp_path / "wb.xlsx"
    full_time = mk("A U", "1", "Software Engineer", 5.0)
    temp = mk("B U", "2", "Temporary Software Engineer", 3.0)
    workbook.write(out, shortlist=[full_time], part_time=[temp], all_postings=[full_time, temp],
                   history_rows=[], changes={"new": [], "closed": []})
    wb = load_workbook(out)
    pt_ws = wb["Part-Time & Temporary"]
    headers = [pt_ws.cell(row=4, column=c).value for c in range(1, pt_ws.max_column + 1)]
    title_col = headers.index("Title") + 1
    assert pt_ws.cell(row=5, column=title_col).value == "Temporary Software Engineer"

    short_ws = wb["Shortlist"]
    headers = [short_ws.cell(row=4, column=c).value for c in range(1, short_ws.max_column + 1)]
    title_col = headers.index("Title") + 1
    assert short_ws.cell(row=5, column=title_col).value == "Software Engineer"


def test_part_time_tab_defaults_to_empty_when_not_passed(tmp_path):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer")
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Part-Time & Temporary"]
    assert ws["A4"].value == "(nothing to show)"


def test_hidden_columns_stay_hidden_but_present(tmp_path):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer", department="CS Dept")
    p.sponsorship_evidence = "may sponsor H-1B"
    p.hard_blockers = ["Security clearance required"]
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]

    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    for col_name, header in (("department", "Department"),
                              ("sponsorship_evidence", "Sponsorship evidence"),
                              ("hard_blockers", "Blockers")):
        assert header in headers, f"{col_name} should still be present, just hidden"
        letter = ws.cell(row=4, column=headers.index(header) + 1).column_letter
        assert ws.column_dimensions[letter].hidden is True, f"{col_name} should be hidden"

    # State is back to being its own visible column -- the Location+State
    # merge was reverted.
    state_letter = ws.cell(row=4, column=headers.index("State") + 1).column_letter
    assert ws.column_dimensions[state_letter].hidden is not True


def test_url_column_is_gone_and_title_carries_the_hyperlink_instead(tmp_path):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer")  # mk() gives it url="https://x/1"
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "URL" not in headers

    title_col = headers.index("Title") + 1
    cell = ws.cell(row=5, column=title_col)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://x/1"
    assert cell.value == "Software Engineer"


def test_summary_shortlist_breakdowns_exclude_negative_scores(tmp_path):
    out = tmp_path / "wb.xlsx"
    good = mk("Good U", "1", "Software Engineer", score=5.0, flag="h1b_possible")
    bad = mk("Bad U", "2", "Software Engineer", score=-1.0, flag="no_sponsorship_any")
    workbook.write(out, shortlist=[good, bad], all_postings=[good, bad],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Summary"]
    values = {ws.cell(row=r, column=c).value
              for r in range(1, ws.max_row + 1) for c in (1, 2)}
    assert "Good U" in values
    assert "Bad U" not in values
    # sponsorship-flag breakdown is filtered the same way
    assert "no_sponsorship_any" not in values
    assert "h1b_possible" in values


def test_summary_all_postings_breakdown_is_not_filtered(tmp_path):
    """All postings by portal/state is a collection-health check, not an
    application-worthiness check -- negative-score rows should still count."""
    out = tmp_path / "wb.xlsx"
    bad = mk("Bad U", "2", "Software Engineer", score=-1.0)
    workbook.write(out, shortlist=[bad], all_postings=[bad],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Summary"]
    values = {ws.cell(row=r, column=c).value
              for r in range(1, ws.max_row + 1) for c in (1, 2)}
    assert "workday" in values     # bad's platform, from mk()


def test_ai_score_key_matches_history_join_scheme():
    assert workbook.ai_score_key("Big State U", "JR123") == "big state u|jr123"
    assert workbook.ai_score_key(None, None) == "|"


def test_ai_score_merges_into_shortlist_by_institution_and_job_id(tmp_path):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer", score=5.0)
    ai_scores = {
        workbook.ai_score_key("A U", "1"): {
            "ai_score": 8, "ai_reason": "Strong match.", "scored_at": "2026-08-08",
        },
    }
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []},
                   ai_scores=ai_scores)
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "AI Score (0-10)" in headers
    assert "AI Score Reason" in headers
    score_col = headers.index("AI Score (0-10)") + 1
    reason_col = headers.index("AI Score Reason") + 1
    assert ws.cell(row=5, column=score_col).value == 8
    assert ws.cell(row=5, column=reason_col).value == "Strong match."


def test_shortlist_is_sorted_by_ai_score_not_regex_score(tmp_path):
    """The regex Score is a cheap pre-filter -- the AI Score is the judgment
    call that actually matters once it exists, so Shortlist row order should
    follow it, even when that means a lower regex-score posting outranks a
    higher one."""
    out = tmp_path / "wb.xlsx"
    low_regex_high_ai = mk("A U", "1", "Low regex, high AI", score=2.0)
    high_regex_low_ai = mk("B U", "2", "High regex, low AI", score=12.0)
    ai_scores = {
        workbook.ai_score_key("A U", "1"): {
            "ai_score": 9, "ai_reason": "Great fit.", "scored_at": "2026-08-08"},
        workbook.ai_score_key("B U", "2"): {
            "ai_score": 1, "ai_reason": "Way underqualified.", "scored_at": "2026-08-08"},
    }
    workbook.write(out, shortlist=[high_regex_low_ai, low_regex_high_ai],
                   all_postings=[high_regex_low_ai, low_regex_high_ai],
                   history_rows=[], changes={"new": [], "closed": []},
                   ai_scores=ai_scores)
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    title_col = headers.index("Title") + 1
    assert ws.cell(row=5, column=title_col).value == "Low regex, high AI"
    assert ws.cell(row=6, column=title_col).value == "High regex, low AI"


def test_shortlist_unscored_postings_sink_below_scored_ones(tmp_path):
    out = tmp_path / "wb.xlsx"
    scored = mk("A U", "1", "Scored", score=10.0)
    unscored = mk("B U", "2", "Unscored", score=12.0)
    ai_scores = {
        workbook.ai_score_key("A U", "1"): {
            "ai_score": 3, "ai_reason": "Real stretch.", "scored_at": "2026-08-08"},
    }
    workbook.write(out, shortlist=[unscored, scored], all_postings=[unscored, scored],
                   history_rows=[], changes={"new": [], "closed": []},
                   ai_scores=ai_scores)
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    title_col = headers.index("Title") + 1
    assert ws.cell(row=5, column=title_col).value == "Scored"
    assert ws.cell(row=6, column=title_col).value == "Unscored"


def test_unmatched_posting_gets_blank_ai_score_not_zero(tmp_path):
    """Blank means 'not yet scored', which is a very different thing from
    an actual AI Score of 0 ('scored, and not recommended'). Uses a non-empty
    ai_scores dict (scoring a DIFFERENT posting) so the AI Score column
    exists on the sheet at all -- an empty {} never adds the column, see
    test_no_ai_scores_argument_at_all_is_a_noop below."""
    out = tmp_path / "wb.xlsx"
    scored = mk("A U", "1", "Software Engineer", score=5.0)
    unmatched = mk("B U", "2", "Software Engineer", score=5.0)
    ai_scores = {workbook.ai_score_key("A U", "1"): {"ai_score": 8, "ai_reason": "x"}}
    workbook.write(out, shortlist=[scored, unmatched], all_postings=[scored, unmatched],
                   history_rows=[], changes={"new": [], "closed": []},
                   ai_scores=ai_scores)
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    score_col = headers.index("AI Score (0-10)") + 1
    values = {ws.cell(row=r, column=score_col).value for r in (5, 6)}
    assert values == {8, None}


def test_no_ai_scores_argument_at_all_is_a_noop(tmp_path):
    """workbook.write() without ai_scores= (the pre-existing call signature)
    should still work -- ai_scores is optional, not required."""
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer", score=5.0)
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "AI Score (0-10)" not in headers    # column doesn't even appear if unused


def test_summary_ai_score_breakdown_buckets_by_exact_value_and_unscored(tmp_path):
    out = tmp_path / "wb.xlsx"
    scored_high = mk("A U", "1", "Software Engineer", score=5.0)
    scored_zero = mk("B U", "2", "Software Engineer", score=3.0)
    unscored = mk("C U", "3", "Software Engineer", score=2.0)
    ai_scores = {
        workbook.ai_score_key("A U", "1"): {"ai_score": 9, "ai_reason": "x"},
        workbook.ai_score_key("B U", "2"): {"ai_score": 0, "ai_reason": "sponsorship hard stop"},
    }
    workbook.write(out, shortlist=[scored_high, scored_zero, unscored],
                   all_postings=[scored_high, scored_zero, unscored],
                   history_rows=[], changes={"new": [], "closed": []},
                   ai_scores=ai_scores)
    ws = load_workbook(out)["Summary"]
    values = {ws.cell(row=r, column=c).value
              for r in range(1, ws.max_row + 1) for c in (1, 2)}
    assert "9" in values
    assert "0 (not recommended)" in values
    assert "(not yet AI-scored)" in values


def test_summary_shortlist_breakdowns_added_for_portal_and_state(tmp_path):
    out = tmp_path / "wb.xlsx"
    good = mk("Good U", "1", "Software Engineer", score=5.0)
    workbook.write(out, shortlist=[good], all_postings=[good],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Summary"]
    headings = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(h and h.startswith("Shortlist by portal") for h in headings)
    assert any(h and h.startswith("Shortlist by state") for h in headings)


def test_summary_shortlist_block_sits_directly_above_its_all_postings_counterpart(tmp_path):
    out = tmp_path / "wb.xlsx"
    good = mk("Good U", "1", "Software Engineer", score=5.0)
    workbook.write(out, shortlist=[good], all_postings=[good],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Summary"]
    headings = [(r, ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value]

    def row_of(prefix):
        return next(r for r, h in headings if h.startswith(prefix))

    assert row_of("Shortlist by portal") < row_of("All postings by portal")
    assert row_of("Shortlist by state") < row_of("All postings by state")


def test_location_and_state_are_separate_columns(tmp_path):
    """Reverted the earlier Location+State merge -- back to two plain columns."""
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer", location="Minneapolis")  # mk() sets state="MN"
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    loc_col = headers.index("Location") + 1
    state_col = headers.index("State") + 1
    assert ws.cell(row=5, column=loc_col).value == "Minneapolis"
    assert ws.cell(row=5, column=state_col).value == "MN"


def test_days_since_posted_column_is_computed_at_export_time(tmp_path, monkeypatch):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer", posted_date=date(2026, 8, 5))
    monkeypatch.setattr(workbook, "date", _FixedDate)
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Days Since Posted" in headers
    col = headers.index("Days Since Posted") + 1
    assert ws.cell(row=5, column=col).value == 3


class _FixedDate(date):
    @classmethod
    def today(cls):
        return date(2026, 8, 8)


def test_description_scraped_column_survives_the_bulky_drop(tmp_path):
    out = tmp_path / "wb.xlsx"
    scraped = mk("A U", "1", "Software Engineer")
    scraped.description_scraped = 1
    unscraped = mk("A U", "2", "Data Analyst")
    unscraped.description_scraped = 0
    workbook.write(out, shortlist=[scraped, unscraped], all_postings=[scraped, unscraped],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Description Verified?" in headers
    col = headers.index("Description Verified?") + 1
    values = {ws.cell(row=r, column=col).value for r in (5, 6)}
    assert values == {1, 0}


def test_description_column_dropped_from_overview_tabs(tmp_path):
    out = tmp_path / "wb.xlsx"
    p = mk("A U", "1", "Software Engineer")
    p.description = "x" * 5000
    workbook.write(out, shortlist=[p], all_postings=[p],
                   history_rows=[], changes={"new": [], "closed": []})
    ws = load_workbook(out)["Shortlist"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Description" not in headers


def test_empty_tabs_render_without_crashing(tmp_path):
    out = tmp_path / "wb.xlsx"
    workbook.write(out, shortlist=[], all_postings=[],
                   history_rows=[], changes={"new": [], "closed": []})
    wb = load_workbook(out)
    assert wb["Changes"]["A4"].value == "(nothing to show)"


def _touch_workbooks(out_dir, stamps):
    for s in stamps:
        (out_dir / f"workbook_{s}.xlsx").write_bytes(b"")


def _row(inst, score, sponsorship="unknown", verified=0):
    return {"institution": inst, "score": score, "sponsorship_flag": sponsorship,
            "description_scraped": verified}


def test_institution_rankings_rewards_several_solid_over_one_outlier():
    rows = [
        _row("Lucky U", 11.5, verified=1),
        _row("Solid State", 7.0, verified=1), _row("Solid State", 6.5, verified=1),
        _row("Solid State", 6.0, verified=1), _row("Solid State", 5.5, verified=1),
    ]
    ranked = workbook.institution_rankings(rows)
    names = [r["institution"] for r in ranked]
    assert names[0] == "Solid State"      # four solid postings beat one lucky 11.5
    solid = next(r for r in ranked if r["institution"] == "Solid State")
    assert solid["postings"] == 4
    assert solid["top3_avg_score"] == round((7.0 + 6.5 + 6.0) / 3, 2)
    assert solid["verified_pct"] == 100


def test_institution_rankings_verified_pct_and_sponsorship_counts():
    rows = [
        _row("Mixed U", 5.0, sponsorship="h1b_possible", verified=1),
        _row("Mixed U", 3.0, sponsorship="no_sponsorship_any", verified=0),
    ]
    ranked = workbook.institution_rankings(rows)
    r = ranked[0]
    assert r["postings"] == 2
    assert r["verified_pct"] == 50
    assert r["positive_sponsorship"] == 1
    assert r["no_sponsorship"] == 1


def test_institution_rankings_sixth_posting_barely_moves_composite():
    five = [_row("Big School", 5.0, verified=1) for _ in range(5)]
    six = five + [_row("Big School", 5.0, verified=1)]
    c5 = workbook.institution_rankings(five)[0]["composite"]
    c6 = workbook.institution_rankings(six)[0]["composite"]
    assert c5 == c6      # sqrt(min(n, 5)) caps the volume bonus at 5


def test_prune_keeps_only_the_newest_n(tmp_path):
    _touch_workbooks(tmp_path, [
        "2026-08-01_120000", "2026-08-02_120000", "2026-08-03_120000",
        "2026-08-04_120000", "2026-08-05_120000", "2026-08-06_120000",
    ])
    deleted = workbook.prune(tmp_path, keep=5)
    remaining = sorted(p.name for p in tmp_path.glob("workbook_*.xlsx"))
    assert len(remaining) == 5
    assert "workbook_2026-08-01_120000.xlsx" not in remaining
    assert len(deleted) == 1


def test_prune_ignores_non_workbook_files(tmp_path):
    _touch_workbooks(tmp_path, ["2026-08-01_120000", "2026-08-02_120000"])
    (tmp_path / "postings.json").write_text("[]")
    workbook.prune(tmp_path, keep=1)
    assert (tmp_path / "postings.json").exists()
    assert len(list(tmp_path.glob("workbook_*.xlsx"))) == 1


def test_prune_keep_zero_or_under_limit_deletes_nothing(tmp_path):
    _touch_workbooks(tmp_path, ["2026-08-01_120000", "2026-08-02_120000"])
    deleted = workbook.prune(tmp_path, keep=5)
    assert deleted == []
    assert len(list(tmp_path.glob("workbook_*.xlsx"))) == 2


def test_prune_dry_run_reports_but_does_not_delete(tmp_path):
    _touch_workbooks(tmp_path, [
        "2026-08-01_120000", "2026-08-02_120000", "2026-08-03_120000",
    ])
    would_delete = workbook.prune(tmp_path, keep=1, dry_run=True)
    assert len(would_delete) == 2
    assert len(list(tmp_path.glob("workbook_*.xlsx"))) == 3   # nothing actually removed
